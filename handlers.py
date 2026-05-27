import logging
import os
import shutil
from datetime import datetime, timedelta
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, BotCommand
from telegram.ext import ContextTypes, ConversationHandler
from sqlalchemy import select, func, update as sql_update, delete
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO

from config import Config
from database import AsyncSessionLocal, is_post_parsed, mark_post_parsed
from models import User, Project, SourceChannel, TargetChannel, ParsedPost, PostQueue, PublishedPost
from scraper import TelegramScraper
from utils import (
    extract_channel_username, calculate_score, clean_caption,
    calculate_next_post_time, get_moscow_time, format_datetime, format_number
)
from backup import BackupService

logger = logging.getLogger(__name__)

# Состояния для ConversationHandler
AWAITING_SOURCE_USERNAME = 1
AWAITING_TARGET_FORWARD = 2
AWAITING_CRITERIA = 3
AWAITING_INTERVAL = 4
AWAITING_PROJECT_NAME = 5
AWAITING_VIEWS = 10
AWAITING_REACTIONS = 11
AWAITING_SIGNATURE = 12
AWAITING_POST_INTERVAL = 13
AWAITING_POST_START_TIME = 14
AWAITING_MEDIA_FILTER = 15
AWAITING_REMOVE_TEXT = 16
AWAITING_BROADCAST_MESSAGE = 17

CURRENT_PROJECT_KEY = "current_project_id"


# ============ ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ============

async def get_current_project(telegram_id: int, context: ContextTypes.DEFAULT_TYPE) -> Project:
    """Получает текущий проект пользователя из кэша контекста, иначе берёт первый активный."""
    project_id = context.user_data.get(CURRENT_PROJECT_KEY)
    
    async with AsyncSessionLocal() as session:
        if project_id:
            result = await session.execute(
                select(Project).where(
                    Project.id == project_id,
                    Project.user_id == telegram_id
                )
            )
            project = result.scalar_one_or_none()
            if project:
                return project
        
        # Если кэш пуст или проект не найден — берём первый активный
        result = await session.execute(
            select(Project).where(
                Project.user_id == telegram_id,
                Project.is_active == True
            ).order_by(Project.id)
        )
        project = result.scalars().first()
        
        if project:
            context.user_data[CURRENT_PROJECT_KEY] = project.id
        
        return project


async def require_project(update: Update, context: ContextTypes.DEFAULT_TYPE) -> Project:
    """Требует наличия проекта. Если нет — отправляет сообщение и возвращает None."""
    telegram_id = update.effective_user.id
    project = await get_current_project(telegram_id, context)
    
    if not project:
        await update.message.reply_text(
            "❌ У вас нет проектов.\n"
            "Создайте первый проект через /my_projects"
        )
        return None
    
    return project


async def is_admin(telegram_id: int) -> bool:
    """Проверяет, является ли пользователь администратором."""
    async with AsyncSessionLocal() as session:
        result = await session.execute(
            select(User).where(User.telegram_id == telegram_id)
        )
        user = result.scalar_one_or_none()
        return user and user.is_admin


async def setup_bot_commands(application):
    """Устанавливает список команд бота в меню."""
    commands = [
        BotCommand("start", "🏠 Главное меню"),
        BotCommand("my_projects", "📁 Мои проекты"),
        BotCommand("add_source", "📥 Добавить источник"),
        BotCommand("add_target", "📤 Добавить целевой канал"),
        BotCommand("my_sources", "📊 Мои источники"),
        BotCommand("my_targets", "🎯 Мои целевые каналы"),
        BotCommand("set_interval", "⏰ Интервал парсинга"),
        BotCommand("status", "📈 Статистика"),
        BotCommand("parse", "🔄 Парсинг сейчас"),
        BotCommand("queue", "📬 Очередь публикации"),
        BotCommand("postnow", "🚀 Опубликовать сейчас"),
        BotCommand("help", "📋 Помощь"),
    ]
    await application.bot.set_my_commands(commands)


async def get_sources_count(project_id: int) -> int:
    """Возвращает количество активных источников в проекте."""
    async with AsyncSessionLocal() as session:
        result = await session.execute(
            select(func.count()).select_from(SourceChannel).where(
                SourceChannel.project_id == project_id,
                SourceChannel.is_active == True
            )
        )
        return result.scalar()


async def get_project_target(project_id: int) -> TargetChannel:
    """Возвращает целевой канал проекта."""
    async with AsyncSessionLocal() as session:
        result = await session.execute(
            select(TargetChannel).where(
                TargetChannel.project_id == project_id,
                TargetChannel.is_active == True
            )
        )
        return result.scalar_one_or_none()


async def get_user_projects_count(telegram_id: int) -> int:
    """Возвращает количество проектов пользователя."""
    async with AsyncSessionLocal() as session:
        result = await session.execute(
            select(func.count()).select_from(Project).where(Project.user_id == telegram_id)
        )
        return result.scalar()


async def send_project_ready_message(update: Update, project_name: str):
    """Отправляет сообщение о готовности проекта к работе."""
    text = (
        f"✅ <b>Проект «{project_name}» готов к работе!</b>\n\n"
        f"📋 Что дальше:\n"
        f"• /set_interval — настроить частоту парсинга\n"
        f"• /parse — запустить первый парсинг\n"
        f"• /status — смотреть статистику\n\n"
        f"🤖 Бот начнёт автоматическую работу согласно настройкам."
    )
    await update.message.reply_text(text, parse_mode="HTML")


# ============ ОСНОВНЫЕ КОМАНДЫ ============

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик команды /start."""
    user = update.effective_user
    is_new_user = False
    
    async with AsyncSessionLocal() as session:
        result = await session.execute(select(User).where(User.telegram_id == user.id))
        db_user = result.scalar_one_or_none()
        
        if not db_user:
            is_new_user = True
            db_user = User(
                telegram_id=user.id,
                username=user.username,
                full_name=user.full_name,
                is_admin=(user.id == Config.ADMIN_ID),
                max_projects=Config.DEFAULT_MAX_PROJECTS,
                max_sources_per_project=Config.DEFAULT_MAX_SOURCES_PER_PROJECT
            )
            session.add(db_user)
            await session.commit()
            logger.info(f"New user: {user.id}")
        
        result = await session.execute(
            select(func.count()).select_from(Project).where(Project.user_id == user.id)
        )
        projects_count = result.scalar()
        has_project = projects_count > 0
    
    if is_new_user and user.id != Config.ADMIN_ID:
        try:
            await context.bot.send_message(
                chat_id=Config.ADMIN_ID,
                text=(
                    f"🆕 <b>Новый пользователь!</b>\n\n"
                    f"👤 {user.full_name or '—'}\n"
                    f"📝 @{user.username or 'нет username'}\n"
                    f"🆔 <code>{user.id}</code>\n"
                    f"📅 {datetime.now().strftime('%d.%m.%Y %H:%M')}"
                ),
                parse_mode="HTML"
            )
        except Exception as e:
            logger.error(f"Failed to notify admin: {e}")
    
    welcome = (
        f"👋 Привет, {user.first_name or 'пользователь'}!\n\n"
        "Я бот для автоматического парсинга и публикации постов из Telegram-каналов.\n\n"
    )
    
    if not has_project:
        welcome += (
            "🚀 Для начала работы создайте первый проект:\n"
            "/my_projects — перейти к проектам\n\n"
        )
    
    welcome += (
        "📋 Основные команды:\n"
        "/my_projects — мои проекты\n"
        "/add_source — добавить источник\n"
        "/add_target — добавить целевой канал\n"
        "/status — статистика\n"
        "/help — все команды"
    )
    
    await update.message.reply_text(welcome)


async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик команды /help."""
    text = (
        "📚 <b>Справка по командам</b>\n\n"
        "<b>Проекты:</b>\n"
        "/my_projects - список ваших проектов\n\n"
        "<b>Источники:</b>\n"
        "/add_source - добавить канал для парсинга\n"
        "/my_sources - список источников\n\n"
        "<b>Целевые каналы:</b>\n"
        "/add_target - добавить канал для публикации\n"
        "/my_targets - список целевых каналов\n\n"
        "<b>Настройки:</b>\n"
        "/set_interval - интервал парсинга\n\n"
        "<b>Управление:</b>\n"
        "/status - общая статистика\n"
        "/project_stats - статистика по проекту\n"
        "/parse - запустить парсинг сейчас\n"
        "/queue - очередь публикации\n"
        "/postnow - опубликовать следующий пост немедленно\n"
    )
    
    if await is_admin(update.effective_user.id):
        text += "\n<b>Админ:</b> /admin"
    
    await update.message.reply_text(text, parse_mode="HTML")


# ============ АДМИН-ПАНЕЛЬ ============

async def admin_panel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показывает админ-панель."""
    if not await is_admin(update.effective_user.id):
        await update.message.reply_text("❌ Нет доступа")
        return
    
    keyboard = [
        [InlineKeyboardButton("👥 Пользователи", callback_data="admin_users_list")],
        [InlineKeyboardButton("💾 Создать бэкап", callback_data="admin_backup_create")],
        [InlineKeyboardButton("📦 Список бэкапов", callback_data="admin_backup_list")],
        [InlineKeyboardButton("📊 Экспорт в Excel", callback_data="admin_export")],
        [InlineKeyboardButton("🔍 Диагностика", callback_data="admin_diagnose")],
        [InlineKeyboardButton("🧹 Очистить очередь", callback_data="admin_clear_queue")],
        [InlineKeyboardButton("🗑️ Очистить failed", callback_data="admin_clear_failed")],
    ]
    
    if update.callback_query:
        await update.callback_query.edit_message_text(
            "👑 <b>Админ-панель</b>\n\nВыберите действие:",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode="HTML"
        )
    else:
        await update.message.reply_text(
            "👑 <b>Админ-панель</b>\n\nВыберите действие:",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode="HTML"
        )


async def admin_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик callback'ов админ-панели."""
    query = update.callback_query
    await query.answer()
    
    if not await is_admin(update.effective_user.id):
        await query.edit_message_text("❌ Нет доступа")
        return
    
    action = query.data
    
    if action == "admin_users_list":
        await show_admin_users(query)
    elif action == "admin_backup_create":
        await create_backup_admin(query)
    elif action == "admin_backup_list":
        await list_backups_admin(query)
    elif action == "admin_export":
        await export_users_excel(query, context)
    elif action == "admin_diagnose":
        await show_diagnose_admin(query)
    elif action == "admin_clear_queue":
        await clear_queue_admin(query)
    elif action == "admin_clear_failed":
        await clear_failed_admin(query)


async def show_admin_users(query):
    """Показывает список пользователей."""
    async with AsyncSessionLocal() as session:
        result = await session.execute(
            select(User).order_by(User.created_at.desc()).limit(20)
        )
        users = result.scalars().all()
    
    text = f"👥 <b>Пользователи ({len(users)}):</b>\n\n"
    for u in users:
        projects_count = await get_user_projects_count(u.telegram_id)
        text += f"• {u.full_name or '—'} (@{u.username or '—'})\n"
        text += f"  🆔 {u.telegram_id} | 📁 {projects_count} проектов\n\n"
    
    keyboard = [[InlineKeyboardButton("◀️ Назад", callback_data="admin_back")]]
    await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode="HTML")


async def create_backup_admin(query):
    """Создаёт бэкап базы данных."""
    await query.edit_message_text("📦 Создаю бэкап...")
    
    backup_service = BackupService()
    backup_path = backup_service.create_backup()
    
    if backup_path:
        try:
            with open(backup_path, 'rb') as f:
                await query.message.reply_document(
                    document=f,
                    filename=os.path.basename(backup_path),
                    caption=f"✅ Бэкап создан\n📅 {datetime.now().strftime('%d.%m.%Y %H:%M')}"
                )
        except Exception as e:
            logger.error(f"Failed to send backup file: {e}")
        
        keyboard = [[InlineKeyboardButton("◀️ Назад", callback_data="admin_back")]]
        await query.edit_message_text(
            f"✅ Бэкап создан и отправлен!\n\n📁 {os.path.basename(backup_path)}",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
    else:
        keyboard = [[InlineKeyboardButton("◀️ Назад", callback_data="admin_back")]]
        await query.edit_message_text(
            "❌ Ошибка создания бэкапа",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )


async def list_backups_admin(query):
    """Показывает список бэкапов."""
    backup_service = BackupService()
    backups = backup_service.list_backups()
    
    if not backups:
        text = "📭 Бэкапов нет"
    else:
        text = "📦 <b>Бэкапы:</b>\n\n"
        for b in backups[:10]:
            text += f"• {b['name']}\n"
            text += f"  📅 {b['created']} | 📦 {b['size_mb']} MB\n\n"
    
    keyboard = [[InlineKeyboardButton("◀️ Назад", callback_data="admin_back")]]
    await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode="HTML")


async def export_users_excel(query, context):
    """Экспортирует список пользователей в Excel."""
    await query.edit_message_text("📊 Формирую отчёт...")
    
    async with AsyncSessionLocal() as session:
        result = await session.execute(select(User).order_by(User.created_at.desc()))
        users = result.scalars().all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Пользователи"
    
    headers = ["Telegram ID", "Username", "Full Name", "Admin", "Projects", "Parsed Today", "Posted Today", "Created At"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    for row, u in enumerate(users, 2):
        projects_count = await get_user_projects_count(u.telegram_id)
        ws.cell(row=row, column=1, value=u.telegram_id)
        ws.cell(row=row, column=2, value=u.username or "")
        ws.cell(row=row, column=3, value=u.full_name or "")
        ws.cell(row=row, column=4, value="Да" if u.is_admin else "Нет")
        ws.cell(row=row, column=5, value=projects_count)
        ws.cell(row=row, column=6, value=u.posts_parsed_today)
        ws.cell(row=row, column=7, value=u.posts_posted_today)
        ws.cell(row=row, column=8, value=u.created_at.strftime("%Y-%m-%d %H:%M") if u.created_at else "")
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    await context.bot.send_document(
        chat_id=query.message.chat_id,
        document=output,
        filename=f"users_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        caption="📊 Экспорт пользователей"
    )
    
    keyboard = [[InlineKeyboardButton("◀️ Назад", callback_data="admin_back")]]
    await query.edit_message_text(
        "✅ Отчёт отправлен!",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )


async def show_diagnose_admin(query):
    """Показывает диагностическую информацию."""
    text = "🔍 <b>Диагностика системы</b>\n\n"
    
    if os.path.exists(Config.DB_PATH):
        size = os.path.getsize(Config.DB_PATH) / (1024 * 1024)
        text += f"📁 БД: {Config.DB_PATH} ({size:.2f} MB)\n"
    else:
        text += f"❌ БД не найдена: {Config.DB_PATH}\n"
    
    text += f"📂 Data: {'✅' if os.path.exists(Config.DATA_DIR) else '❌'}\n"
    text += f"📂 Temp: {'✅' if os.path.exists(Config.TEMP_DIR) else '❌'}\n"
    text += f"📂 Backups: {'✅' if os.path.exists(Config.BACKUP_DIR) else '❌'}\n"
    
    async with AsyncSessionLocal() as session:
        result = await session.execute(select(func.count()).select_from(User))
        users_count = result.scalar()
        result = await session.execute(select(func.count()).select_from(Project))
        projects_count = result.scalar()
        result = await session.execute(select(func.count()).select_from(PostQueue).where(PostQueue.status == "pending"))
        pending = result.scalar()
    
    text += f"\n👥 Пользователей: {users_count}\n"
    text += f"📁 Проектов: {projects_count}\n"
    text += f"📬 В очереди: {pending}\n"
    
    keyboard = [[InlineKeyboardButton("◀️ Назад", callback_data="admin_back")]]
    await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode="HTML")


async def clear_queue_admin(query):
    """Очищает очередь публикации."""
    async with AsyncSessionLocal() as session:
        result = await session.execute(select(PostQueue).where(PostQueue.status == "pending"))
        items = result.scalars().all()
        deleted = len(items)
        for item in items:
            await session.delete(item)
        await session.commit()
    
    keyboard = [[InlineKeyboardButton("◀️ Назад", callback_data="admin_back")]]
    await query.edit_message_text(
        f"✅ Удалено {deleted} постов из очереди",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )


async def clear_failed_admin(query):
    """Очищает failed посты из очереди."""
    async with AsyncSessionLocal() as session:
        result = await session.execute(select(PostQueue).where(PostQueue.status == "failed"))
        items = result.scalars().all()
        deleted = len(items)
        for item in items:
            await session.delete(item)
        await session.commit()
    
    keyboard = [[InlineKeyboardButton("◀️ Назад", callback_data="admin_back")]]
    await query.edit_message_text(
        f"✅ Удалено {deleted} failed постов",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )


async def admin_back_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Возвращает в главное меню админ-панели."""
    query = update.callback_query
    await query.answer()
    
    keyboard = [
        [InlineKeyboardButton("👥 Пользователи", callback_data="admin_users_list")],
        [InlineKeyboardButton("💾 Создать бэкап", callback_data="admin_backup_create")],
        [InlineKeyboardButton("📦 Список бэкапов", callback_data="admin_backup_list")],
        [InlineKeyboardButton("📊 Экспорт в Excel", callback_data="admin_export")],
        [InlineKeyboardButton("🔍 Диагностика", callback_data="admin_diagnose")],
        [InlineKeyboardButton("🧹 Очистить очередь", callback_data="admin_clear_queue")],
        [InlineKeyboardButton("🗑️ Очистить failed", callback_data="admin_clear_failed")],
    ]
    
    await query.edit_message_text(
        "👑 <b>Админ-панель</b>\n\nВыберите действие:",
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode="HTML"
    )


# ============ УПРАВЛЕНИЕ ПРОЕКТАМИ ============

async def my_projects(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показывает список проектов пользователя."""
    telegram_id = update.effective_user.id
    current_project = await get_current_project(telegram_id, context)
    
    async with AsyncSessionLocal() as session:
        result = await session.execute(select(User).where(User.telegram_id == telegram_id))
        user = result.scalar_one()
        
        result = await session.execute(
            select(Project).where(Project.user_id == telegram_id).order_by(Project.id)
        )
        projects = result.scalars().all()
    
    if not projects:
        keyboard = [[InlineKeyboardButton("➕ Создать проект", callback_data="create_project")]]
        await update.message.reply_text(
            "📭 У вас пока нет проектов.\n\n"
            "Проект — это связка из целевого канала и источников.\n"
            "Например: «Мемасы», «Книги», «Кино»",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        return
    
    text = f"📁 <b>Ваши проекты</b> ({len(projects)} / {user.max_projects})\n\n"
    keyboard = []
    
    for p in projects:
        sources_count = await get_sources_count(p.id)
        target = await get_project_target(p.id)
        
        current_icon = "👉 " if current_project and p.id == current_project.id else ""
        
        text += f"{current_icon}<b>{p.name}</b>\n"
        text += f"   📥 Источников: {sources_count}\n"
        text += f"   📤 Цель: {target.channel_title if target else 'не задан'}\n"
        text += f"   📊 Сегодня: {p.posts_parsed_today} / {p.posts_posted_today}\n\n"
        
        if not current_project or p.id != current_project.id:
            keyboard.append([InlineKeyboardButton(f"✅ Выбрать «{p.name}»", callback_data=f"select_project_{p.id}")])
        
        keyboard.append([
            InlineKeyboardButton(f"📊 Статистика", callback_data=f"stats_project_{p.id}"),
            InlineKeyboardButton(f"❌ Удалить", callback_data=f"delete_project_{p.id}")
        ])
    
    if len(projects) < user.max_projects:
        keyboard.append([InlineKeyboardButton("➕ Создать новый проект", callback_data="create_project")])
    
    await update.message.reply_text(
        text,
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode="HTML"
    )


async def projects_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик callback'ов управления проектами."""
    query = update.callback_query
    await query.answer()
    telegram_id = update.effective_user.id
    data = query.data
    
    if data == "create_project":
        await query.edit_message_text("📁 Введите название для нового проекта:")
        context.user_data['awaiting_project_name'] = True
        return
    
    if data.startswith("select_project_"):
        project_id = int(data.replace("select_project_", ""))
        async with AsyncSessionLocal() as session:
            result = await session.execute(
                select(Project).where(Project.id == project_id, Project.user_id == telegram_id)
            )
            project = result.scalar_one_or_none()
        
        if project:
            context.user_data[CURRENT_PROJECT_KEY] = project.id
            await query.edit_message_text(f"✅ Выбран проект «{project.name}»")
    
    elif data.startswith("delete_project_"):
        project_id = int(data.replace("delete_project_", ""))
        keyboard = [
            [InlineKeyboardButton("✅ Да, удалить", callback_data=f"confirm_delete_{project_id}")],
            [InlineKeyboardButton("❌ Отмена", callback_data="cancel_delete")],
        ]
        await query.edit_message_text(
            "⚠️ Удалить проект? Все источники и настройки будут потеряны.",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
    
    elif data.startswith("confirm_delete_"):
        project_id = int(data.replace("confirm_delete_", ""))
        async with AsyncSessionLocal() as session:
            await session.execute(delete(SourceChannel).where(SourceChannel.project_id == project_id))
            await session.execute(delete(TargetChannel).where(TargetChannel.project_id == project_id))
            await session.execute(delete(PostQueue).where(PostQueue.project_id == project_id))
            await session.execute(delete(Project).where(Project.id == project_id))
            await session.commit()
        
        if context.user_data.get(CURRENT_PROJECT_KEY) == project_id:
            context.user_data.pop(CURRENT_PROJECT_KEY, None)
        
        await query.edit_message_text("✅ Проект удалён")
    
    elif data == "cancel_delete":
        await query.edit_message_text("❌ Удаление отменено")
    
    elif data.startswith("stats_project_"):
        project_id = int(data.replace("stats_project_", ""))
        await show_project_stats(query, project_id)


async def handle_project_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает ввод названия нового проекта."""
    if not context.user_data.get('awaiting_project_name'):
        return
    
    name = update.message.text.strip()
    telegram_id = update.effective_user.id
    
    if len(name) < 2 or len(name) > 50:
        await update.message.reply_text("❌ Название должно быть от 2 до 50 символов.")
        return
    
    async with AsyncSessionLocal() as session:
        project = Project(
            user_id=telegram_id,
            name=name,
            check_interval_minutes=Config.DEFAULT_CHECK_INTERVAL,
            post_interval_hours=Config.DEFAULT_POST_INTERVAL_HOURS,
            active_hours_start=Config.DEFAULT_ACTIVE_HOURS_START,
            active_hours_end=Config.DEFAULT_ACTIVE_HOURS_END
        )
        session.add(project)
        await session.commit()
        context.user_data[CURRENT_PROJECT_KEY] = project.id
    
    context.user_data['awaiting_project_name'] = False
    
    await update.message.reply_text(
        f"✅ Проект «{name}» создан!\n\n"
        f"Теперь добавьте:\n"
        f"• /add_target — целевой канал\n"
        f"• /add_source — каналы-источники"
    )


async def show_project_stats(query, project_id: int):
    """Показывает статистику конкретного проекта."""
    async with AsyncSessionLocal() as session:
        result = await session.execute(select(Project).where(Project.id == project_id))
        project = result.scalar_one()
    
    sources_count = await get_sources_count(project_id)
    target = await get_project_target(project_id)
    
    async with AsyncSessionLocal() as session:
        result = await session.execute(
            select(PostQueue).where(PostQueue.project_id == project_id, PostQueue.status == "pending")
        )
        pending = len(result.scalars().all())
    
    text = (
        f"📊 <b>Статистика «{project.name}»</b>\n\n"
        f"📥 Источников: {sources_count}\n"
        f"📤 Цель: {target.channel_title if target else 'не задан'}\n"
        f"⏰ Интервал: {project.check_interval_minutes} мин\n"
        f"📈 Сегодня: {project.posts_parsed_today} / {project.posts_posted_today}\n"
        f"📬 В очереди: {pending}"
    )
    
    keyboard = [[InlineKeyboardButton("◀️ Назад к проектам", callback_data="back_to_projects")]]
    await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode="HTML")


async def back_to_projects_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Возвращает к списку проектов."""
    query = update.callback_query
    await query.answer()
    await my_projects(update, context)


# ============ ИСТОЧНИКИ ============

async def add_source_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Начинает добавление источника."""
    telegram_id = update.effective_user.id
    project = await require_project(update, context)
    
    if not project:
        return ConversationHandler.END
    
    async with AsyncSessionLocal() as session:
        result = await session.execute(select(User).where(User.telegram_id == telegram_id))
        user = result.scalar_one()
        
        count = await get_sources_count(project.id)
        if count >= user.max_sources_per_project:
            await update.message.reply_text(
                f"❌ Достигнут лимит источников ({user.max_sources_per_project})."
            )
            return ConversationHandler.END
    
    context.user_data['temp_project_id'] = project.id
    context.user_data['temp_project_name'] = project.name
    
    await update.message.reply_text(
        f"📥 Добавление источника в «{project.name}»\n\n"
        "Отправьте username канала (@name) или ссылку:\n"
        "• @durov\n"
        "• https://t.me/durov"
    )
    return AWAITING_SOURCE_USERNAME


async def add_source_username(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает username источника."""
    username = extract_channel_username(update.message.text)
    if not username:
        await update.message.reply_text("❌ Не удалось распознать username.")
        return AWAITING_SOURCE_USERNAME
    
    async with TelegramScraper() as scraper:
        info = await scraper.get_channel_info(username)
    
    if not info:
        await update.message.reply_text("❌ Канал не найден или не публичный.")
        return AWAITING_SOURCE_USERNAME
    
    context.user_data['temp_source'] = {
        'username': username,
        'title': info['title'],
        'project_id': context.user_data.get('temp_project_id'),
        'project_name': context.user_data.get('temp_project_name')
    }
    
    keyboard = [
        [InlineKeyboardButton("🎯 Свои критерии", callback_data="criteria_custom")],
        [InlineKeyboardButton("👁 1000+ просмотров", callback_data="criteria_views")],
        [InlineKeyboardButton("❤️ 50+ реакций", callback_data="criteria_reactions")],
        [InlineKeyboardButton("👁+❤️ 500+ и 25+", callback_data="criteria_both")],
        [InlineKeyboardButton("⚡ Без критериев", callback_data="criteria_none")],
    ]
    
    await update.message.reply_text(
        f"✅ Канал: @{username}\n"
        f"📝 Название: {info['title']}\n\n"
        f"Выберите критерии:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return AWAITING_CRITERIA


async def add_source_criteria(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает выбор критериев для источника."""
    query = update.callback_query
    await query.answer()
    
    choice = query.data.replace("criteria_", "")
    temp = context.user_data.get('temp_source')
    
    if not temp:
        await query.edit_message_text("❌ Ошибка: данные не найдены.")
        return ConversationHandler.END
    
    if choice == "custom":
        await query.edit_message_text(
            "📊 <b>Настройка критериев</b>\n\n"
            "Введите минимальное количество просмотров (0 = не учитывать):",
            parse_mode="HTML"
        )
        context.user_data['awaiting_criteria'] = 'views'
        return AWAITING_VIEWS
    else:
        criteria = {
            "views": {"min_views": 1000},
            "reactions": {"min_reactions": 50},
            "both": {"min_views": 500, "min_reactions": 25},
            "none": {}
        }.get(choice, {})
        
        await save_source_with_criteria(query, context, temp, criteria)
        return ConversationHandler.END


async def criteria_views_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает ввод минимальных просмотров."""
    try:
        views = int(update.message.text.strip())
        if views < 0:
            raise ValueError
    except:
        await update.message.reply_text("❌ Введите целое число (0 = не учитывать):")
        return AWAITING_VIEWS
    
    context.user_data['temp_criteria_views'] = views
    await update.message.reply_text("📊 Введите минимальное количество реакций (0 = не учитывать):")
    return AWAITING_REACTIONS


async def criteria_reactions_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает ввод минимальных реакций."""
    try:
        reactions = int(update.message.text.strip())
        if reactions < 0:
            raise ValueError
    except:
        await update.message.reply_text("❌ Введите целое число (0 = не учитывать):")
        return AWAITING_REACTIONS
    
    views = context.user_data.get('temp_criteria_views', 0)
    criteria = {}
    if views > 0:
        criteria['min_views'] = views
    if reactions > 0:
        criteria['min_reactions'] = reactions
    
    temp = context.user_data.get('temp_source')
    await save_source_with_criteria(update, context, temp, criteria)
    
    context.user_data.pop('temp_criteria_views', None)
    context.user_data.pop('awaiting_criteria', None)
    return ConversationHandler.END


async def save_source_with_criteria(target, context, temp: dict, criteria: dict):
    """Сохраняет источник с критериями в базу данных."""
    async with AsyncSessionLocal() as session:
        result = await session.execute(
            select(SourceChannel).where(
                SourceChannel.project_id == temp['project_id'],
                SourceChannel.channel_username == temp['username']
            )
        )
        if result.scalar_one_or_none():
            text = f"⚠️ Канал @{temp['username']} уже добавлен."
            if hasattr(target, 'edit_message_text'):
                await target.edit_message_text(text)
            else:
                await target.reply_text(text)
            return
        
        channel = SourceChannel(
            project_id=temp['project_id'],
            channel_username=temp['username'],
            channel_title=temp['title'],
            criteria=criteria
        )
        session.add(channel)
        await session.commit()
    
    criteria_text = []
    if criteria.get('min_views'):
        criteria_text.append(f"👁 ≥{criteria['min_views']}")
    if criteria.get('min_reactions'):
        criteria_text.append(f"❤️ ≥{criteria['min_reactions']}")
    criteria_str = ", ".join(criteria_text) if criteria_text else "без критериев"
    
    text = f"✅ Канал @{temp['username']} добавлен!\n📋 Критерии: {criteria_str}"
    
    if hasattr(target, 'edit_message_text'):
        await target.edit_message_text(text)
    else:
        await target.reply_text(text)
    
    project_id = temp['project_id']
    project_name = temp['project_name']
    
    context.user_data.pop('temp_source', None)
    context.user_data.pop('temp_project_id', None)
    context.user_data.pop('temp_project_name', None)
    
    sources_count = await get_sources_count(project_id)
    target_channel = await get_project_target(project_id)
    if sources_count == 1 and target_channel:
        if hasattr(target, 'message'):
            await send_project_ready_message(target, project_name)
        else:
            await target.message.reply_text(
                f"✅ <b>Проект «{project_name}» готов к работе!</b>\n\n"
                f"• /set_interval — настроить частоту\n"
                f"• /parse — запустить парсинг",
                parse_mode="HTML"
            )


async def my_sources(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показывает список источников текущего проекта."""
    telegram_id = update.effective_user.id
    project = await require_project(update, context)
    
    if not project:
        return
    
    async with AsyncSessionLocal() as session:
        result = await session.execute(
            select(SourceChannel)
            .where(SourceChannel.project_id == project.id)
            .order_by(SourceChannel.added_at.desc())
        )
        sources = result.scalars().all()
        
        result = await session.execute(select(User).where(User.telegram_id == telegram_id))
        user = result.scalar_one()
    
    if not sources:
        await update.message.reply_text(
            f"📭 В проекте «{project.name}» нет источников.\n"
            f"Добавьте: /add_source"
        )
        return
    
    text = f"📥 <b>Источники «{project.name}»</b> ({len(sources)} / {user.max_sources_per_project})\n\n"
    keyboard = []
    
    for src in sources:
        criteria_text = []
        if src.criteria:
            if "min_views" in src.criteria:
                criteria_text.append(f"👁 ≥{src.criteria['min_views']}")
            if "min_reactions" in src.criteria:
                criteria_text.append(f"❤️ ≥{src.criteria['min_reactions']}")
        criteria_str = ", ".join(criteria_text) if criteria_text else "без критериев"
        
        text += f"• @{src.channel_username}\n"
        text += f"  📊 {criteria_str}\n\n"
        
        keyboard.append([
            InlineKeyboardButton(f"❌ Удалить @{src.channel_username}", callback_data=f"del_source_{src.id}")
        ])
    
    await update.message.reply_text(
        text,
        reply_markup=InlineKeyboardMarkup(keyboard) if keyboard else None,
        parse_mode="HTML"
    )


async def delete_source_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Удаляет источник по callback."""
    query = update.callback_query
    await query.answer()
    
    source_id = int(query.data.replace("del_source_", ""))
    
    async with AsyncSessionLocal() as session:
        await session.execute(delete(SourceChannel).where(SourceChannel.id == source_id))
        await session.commit()
    
    await query.edit_message_text("✅ Источник удалён")


# ============ ЦЕЛЕВЫЕ КАНАЛЫ ============

async def add_target_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Добавление целевого канала. Проверяет текущий проект и другие проекты."""
    telegram_id = update.effective_user.id
    project = await require_project(update, context)
    
    if not project:
        return ConversationHandler.END
    
    # Проверяем, есть ли target в ТЕКУЩЕМ проекте
    target = await get_project_target(project.id)
    if target:
        await update.message.reply_text(
            f"⚠️ В текущем проекте «{project.name}» уже есть целевой канал:\n"
            f"📤 {target.channel_title}\n\n"
            f"Чтобы добавить новый, сначала удалите текущий через /my_targets\n"
            f"Или переключитесь на другой проект в /my_projects"
        )
        return ConversationHandler.END
    
    # Проверяем, есть ли target в ДРУГИХ проектах пользователя
    async with AsyncSessionLocal() as session:
        result = await session.execute(
            select(Project).where(
                Project.user_id == telegram_id,
                Project.id != project.id,
                Project.is_active == True
            )
        )
        other_projects = result.scalars().all()
    
    other_with_target = []
    for p in other_projects:
        t = await get_project_target(p.id)
        if t:
            other_with_target.append((p, t))
    
    if other_with_target:
        # Показываем, что target уже есть в других проектах
        text = f"📁 Текущий проект: «{project.name}»\n\n"
        text += "ℹ️ Целевой канал уже существует в других проектах:\n"
        for p, t in other_with_target:
            text += f"• Проект «{p.name}» → {t.channel_title}\n"
        text += f"\nДобавляем новый целевой канал в проект «{project.name}». Продолжить?"
        
        keyboard = []
        for p, t in other_with_target:
            keyboard.append([InlineKeyboardButton(
                f"🔄 Переключиться на «{p.name}»", 
                callback_data=f"select_project_{p.id}"
            )])
        keyboard.append([InlineKeyboardButton(
            f"✅ Да, добавить в «{project.name}»", 
            callback_data="add_target_continue"
        )])
        
        await update.message.reply_text(
            text, 
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        return ConversationHandler.END
    
    # Всё чисто — продолжаем добавление
    context.user_data['temp_project_id'] = project.id
    context.user_data['temp_project_name'] = project.name
    
    me = await context.bot.get_me()
    await update.message.reply_text(
        f"🎯 Добавление целевого канала в «{project.name}»\n\n"
        f"1. Добавьте @{me.username} в админы канала\n"
        f"2. Выдайте права на публикацию\n"
        f"3. Перешлите сюда любое сообщение из канала"
    )
    return AWAITING_TARGET_FORWARD


async def add_target_continue_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Callback для продолжения добавления target после предупреждения."""
    query = update.callback_query
    await query.answer()
    
    telegram_id = update.effective_user.id
    project = await require_project(update, context)
    
    if not project:
        return
    
    # Дополнительно проверяем, что target всё ещё не добавлен
    target = await get_project_target(project.id)
    if target:
        await query.edit_message_text(
            f"⚠️ В проекте «{project.name}» уже есть целевой канал: {target.channel_title}\n"
            f"Удалите его через /my_targets"
        )
        return
    
    context.user_data['temp_project_id'] = project.id
    context.user_data['temp_project_name'] = project.name
    
    me = await context.bot.get_me()
    await query.edit_message_text(
        f"🎯 Добавление целевого канала в «{project.name}»\n\n"
        f"1. Добавьте @{me.username} в админы канала\n"
        f"2. Выдайте права на публикацию\n"
        f"3. Перешлите сюда любое сообщение из канала"
    )
    return AWAITING_TARGET_FORWARD


async def add_target_forward(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает пересланное сообщение из целевого канала."""
    msg = update.message
    
    if not msg.forward_from_chat or msg.forward_from_chat.type != 'channel':
        await update.message.reply_text("❌ Перешлите сообщение из канала.")
        return AWAITING_TARGET_FORWARD
    
    chat = msg.forward_from_chat
    project_id = context.user_data.get('temp_project_id')
    project_name = context.user_data.get('temp_project_name')
    
    try:
        test_msg = await context.bot.send_message(chat.id, "🔧 Проверка прав...")
        await test_msg.delete()
    except Exception as e:
        await update.message.reply_text("❌ Бот не имеет прав администратора.")
        return AWAITING_TARGET_FORWARD
    
    async with AsyncSessionLocal() as session:
        channel = TargetChannel(
            project_id=project_id,
            channel_id=chat.id,
            channel_username=chat.username,
            channel_title=chat.title
        )
        session.add(channel)
        await session.commit()
    
    await update.message.reply_text(
        f"✅ Целевой канал добавлен!\n"
        f"📝 {chat.title}"
    )
    
    context.user_data.pop('temp_project_id', None)
    context.user_data.pop('temp_project_name', None)
    
    sources_count = await get_sources_count(project_id)
    if sources_count > 0:
        await send_project_ready_message(update, project_name)
    
    return ConversationHandler.END


async def my_targets(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показывает целевой канал текущего проекта."""
    project = await require_project(update, context)
    
    if not project:
        return
    
    target = await get_project_target(project.id)
    
    if not target:
        await update.message.reply_text(
            f"📭 В проекте «{project.name}» нет целевого канала.\n"
            f"Добавьте: /add_target"
        )
        return
    
    text = f"🎯 <b>Целевой канал «{project.name}»</b>\n\n"
    text += f"📝 {target.channel_title}\n"
    if target.channel_username:
        text += f"🔗 @{target.channel_username}\n"
    text += f"🆔 {target.channel_id}\n"
    
    keyboard = [[InlineKeyboardButton("❌ Удалить", callback_data=f"del_target_{target.id}")]]
    
    await update.message.reply_text(
        text,
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode="HTML"
    )


async def delete_target_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Удаляет целевой канал по callback."""
    query = update.callback_query
    await query.answer()
    
    target_id = int(query.data.replace("del_target_", ""))
    
    async with AsyncSessionLocal() as session:
        await session.execute(delete(TargetChannel).where(TargetChannel.id == target_id))
        await session.commit()
    
    await query.edit_message_text("✅ Целевой канал удалён")


# ============ НАСТРОЙКИ ============

async def set_interval_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Начинает настройку интервала парсинга."""
    project = await require_project(update, context)
    
    if not project:
        return ConversationHandler.END
    
    context.user_data['temp_project_id'] = project.id
    
    keyboard = [
        [InlineKeyboardButton("🕐 30 минут", callback_data="interval_30")],
        [InlineKeyboardButton("🕑 1 час", callback_data="interval_60")],
        [InlineKeyboardButton("🕒 2 часа", callback_data="interval_120")],
        [InlineKeyboardButton("🕓 3 часа", callback_data="interval_180")],
        [InlineKeyboardButton("🕔 6 часов", callback_data="interval_360")],
        [InlineKeyboardButton("🕕 12 часов", callback_data="interval_720")],
    ]
    
    await update.message.reply_text(
        f"⏰ <b>Интервал парсинга</b>\n\n"
        f"Проект: {project.name}\n"
        f"Текущий: {project.check_interval_minutes} мин\n\n"
        f"Выберите новый интервал:",
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode="HTML"
    )
    return AWAITING_INTERVAL


async def set_interval_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает выбор интервала парсинга."""
    query = update.callback_query
    await query.answer()
    
    interval = int(query.data.replace("interval_", ""))
    project_id = context.user_data.get('temp_project_id')
    
    async with AsyncSessionLocal() as session:
        await session.execute(
            sql_update(Project)
            .where(Project.id == project_id)
            .values(check_interval_minutes=interval)
        )
        await session.commit()
    
    await query.edit_message_text(f"✅ Интервал парсинга: {interval} минут")
    context.user_data.pop('temp_project_id', None)
    return ConversationHandler.END


async def set_post_interval_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Начинает настройку интервала публикации."""
    project = await require_project(update, context)
    
    if not project:
        return ConversationHandler.END
    
    context.user_data['temp_project_id'] = project.id
    
    keyboard = []
    hours_options = [1, 2, 3, 4, 6, 8, 12, 24]
    row = []
    for i, h in enumerate(hours_options):
        row.append(InlineKeyboardButton(f"{h}ч", callback_data=f"post_{h}"))
        if len(row) == 4:
            keyboard.append(row)
            row = []
    if row:
        keyboard.append(row)
    
    await update.message.reply_text(
        f"📅 <b>Интервал публикации</b>\n\n"
        f"Проект: {project.name}\n"
        f"Текущий: каждые {project.post_interval_hours} ч\n\n"
        f"Выберите новый интервал:",
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode="HTML"
    )
    return AWAITING_POST_INTERVAL


async def set_post_interval_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает выбор интервала публикации."""
    query = update.callback_query
    await query.answer()
    
    hours = int(query.data.replace("post_", ""))
    context.user_data['temp_post_interval'] = hours
    
    keyboard = [
        [InlineKeyboardButton("🌅 Утро (08:00-20:00)", callback_data="starttime_morning")],
        [InlineKeyboardButton("🌙 Весь день (00:00-23:59)", callback_data="starttime_allday")],
        [InlineKeyboardButton("🎯 Своё время", callback_data="starttime_custom")],
    ]
    
    await query.edit_message_text(
        f"⏰ <b>Время публикации</b>\n\n"
        f"Интервал: каждые {hours} ч\n\n"
        f"Выберите режим:",
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode="HTML"
    )
    return AWAITING_POST_START_TIME


async def set_post_start_time_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает выбор времени начала публикации."""
    query = update.callback_query
    await query.answer()
    
    choice = query.data.replace("starttime_", "")
    hours = context.user_data.get('temp_post_interval', 2)
    project_id = context.user_data.get('temp_project_id')
    
    if choice == "morning":
        start, end = 8, 20
    elif choice == "allday":
        start, end = 0, 23
    else:
        start, end = 8, 20  # default
    
    async with AsyncSessionLocal() as session:
        await session.execute(
            sql_update(Project)
            .where(Project.id == project_id)
            .values(
                post_interval_hours=hours,
                active_hours_start=start,
                active_hours_end=end
            )
        )
        await session.commit()
    
    time_desc = f"{start}:00 - {end}:00" if choice != "allday" else "круглосуточно"
    
    await query.edit_message_text(
        f"✅ Интервал публикации: каждые {hours} ч\n"
        f"🕐 Время: {time_desc}"
    )
    
    context.user_data.pop('temp_project_id', None)
    context.user_data.pop('temp_post_interval', None)
    return ConversationHandler.END


async def set_signature_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Начинает настройку подписи."""
    project = await require_project(update, context)
    
    if not project:
        return ConversationHandler.END
    
    context.user_data['temp_project_id'] = project.id
    
    await update.message.reply_text(
        f"✍️ <b>Подпись к постам</b>\n\n"
        f"Проект: {project.name}\n"
        f"Текущая: {project.signature or 'нет'}\n\n"
        f"Отправьте текст подписи (или '-' чтобы убрать):",
        parse_mode="HTML"
    )
    return AWAITING_SIGNATURE


async def set_signature_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает ввод подписи."""
    signature = update.message.text.strip()
    project_id = context.user_data.get('temp_project_id')
    
    if signature == '-':
        signature = None
    
    async with AsyncSessionLocal() as session:
        await session.execute(
            sql_update(Project)
            .where(Project.id == project_id)
            .values(signature=signature)
        )
        await session.commit()
    
    await update.message.reply_text(
        f"✅ Подпись {'удалена' if not signature else 'установлена'}: {signature or ''}"
    )
    
    context.user_data.pop('temp_project_id', None)
    return ConversationHandler.END


# ============ СТАТИСТИКА ============

async def status(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показывает общую статистику пользователя."""
    telegram_id = update.effective_user.id
    
    async with AsyncSessionLocal() as session:
        result = await session.execute(select(User).where(User.telegram_id == telegram_id))
        user = result.scalar_one()
        
        result = await session.execute(select(Project).where(Project.user_id == telegram_id))
        projects = result.scalars().all()
        
        total_parsed = sum(p.posts_parsed_today for p in projects)
        total_posted = sum(p.posts_posted_today for p in projects)
    
    text = (
        f"📊 <b>Общая статистика</b>\n\n"
        f"📁 Проектов: {len(projects)} / {user.max_projects}\n"
        f"📈 За сегодня:\n"
        f"• Спарсено: {total_parsed}\n"
        f"• Опубликовано: {total_posted}\n\n"
        f"/my_projects — статистика по проектам"
    )
    
    await update.message.reply_text(text, parse_mode="HTML")


async def project_stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показывает статистику текущего проекта."""
    project = await require_project(update, context)
    
    if not project:
        return
    
    sources_count = await get_sources_count(project.id)
    target = await get_project_target(project.id)
    
    async with AsyncSessionLocal() as session:
        result = await session.execute(
            select(PostQueue).where(PostQueue.project_id == project.id, PostQueue.status == "pending")
        )
        pending = len(result.scalars().all())
    
    text = (
        f"📊 <b>Статистика «{project.name}»</b>\n\n"
        f"📥 Источников: {sources_count}\n"
        f"📤 Цель: {target.channel_title if target else 'не задан'}\n"
        f"⏰ Интервал: {project.check_interval_minutes} мин\n"
        f"📈 Сегодня: {project.posts_parsed_today} / {project.posts_posted_today}\n"
        f"📬 В очереди: {pending}"
    )
    
    await update.message.reply_text(text, parse_mode="HTML")


# ============ ПАРСИНГ И ОЧЕРЕДЬ ============

async def parse_now(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Запускает парсинг вручную."""
    project = await require_project(update, context)
    
    if not project:
        return
    
    target = await get_project_target(project.id)
    if not target:
        await update.message.reply_text("❌ Сначала добавьте целевой канал: /add_target")
        return
    
    sources_count = await get_sources_count(project.id)
    if sources_count == 0:
        await update.message.reply_text("❌ Сначала добавьте источники: /add_source")
        return
    
    msg = await update.message.reply_text(f"🔄 Парсинг «{project.name}»...")
    
    scheduler = context.application.bot_data.get('scheduler')
    if scheduler:
        await scheduler._process_project(project)
        
        async with AsyncSessionLocal() as session:
            result = await session.execute(select(Project).where(Project.id == project.id))
            updated = result.scalar_one()
        
        await msg.edit_text(
            f"✅ Парсинг завершён!\n\n"
            f"📊 Спарсено сегодня: {updated.posts_parsed_today}\n"
            f"📤 В очереди: /queue"
        )
    else:
        await msg.edit_text("❌ Планировщик не найден")


async def queue_status(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показывает очередь публикации."""
    project = await require_project(update, context)
    
    if not project:
        return
    
    async with AsyncSessionLocal() as session:
        result = await session.execute(
            select(PostQueue).where(
                PostQueue.project_id == project.id
            ).order_by(PostQueue.scheduled_time).limit(15)
        )
        items = result.scalars().all()
    
    if not items:
        await update.message.reply_text("📭 Очередь публикации пуста")
        return
    
    text = f"📬 <b>Очередь публикации «{project.name}»</b>\n\n"
    text += f"⏰ Интервал: каждые {project.post_interval_hours} ч\n\n"
    
    MSK_OFFSET = timedelta(hours=3)
    
    for item in items:
        post_data = item.post_data
        status_icon = {"pending": "⏳", "published": "✅", "failed": "❌"}.get(item.status, "❓")
        
        scheduled_msk = item.scheduled_time + MSK_OFFSET
        
        text += f"{status_icon} {scheduled_msk.strftime('%d.%m.%Y %H:%M')} МСК\n"
        text += f"   📡 @{post_data.get('source_username', '?')}\n"
        text += f"   👁 {format_number(post_data.get('views', 0))} | ❤️ {format_number(post_data.get('reactions', 0))}\n\n"
    
    await update.message.reply_text(text, parse_mode="HTML")


async def post_now(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Опубликовать следующий пост из очереди немедленно (админ)."""
    project = await require_project(update, context)
    
    if not project:
        return
    
    if not await is_admin(update.effective_user.id):
        await update.message.reply_text("❌ Только админ может принудительно публиковать посты")
        return
    
    async with AsyncSessionLocal() as session:
        result = await session.execute(
            select(PostQueue).where(
                PostQueue.project_id == project.id,
                PostQueue.status == "pending"
            ).order_by(PostQueue.scheduled_time).limit(1)
        )
        queue_item = result.scalar_one_or_none()
        
        if not queue_item:
            await update.message.reply_text("📭 Нет постов в очереди для публикации")
            return
        
        poster = context.application.bot_data.get('poster')
        if not poster:
            await update.message.reply_text("❌ Сервис публикации не найден")
            return
        
        msg = await update.message.reply_text("🚀 Публикую пост...")
        
        success = await poster.publish_post(queue_item)
        
        if success:
            result = await session.execute(select(Project).where(Project.id == project.id))
            db_project = result.scalar_one()
            db_project.posts_posted_today += 1
            await session.commit()
            
            post_data = queue_item.post_data
            await msg.edit_text(
                f"✅ Пост опубликован!\n\n"
                f"📡 @{post_data.get('source_username', '?')}\n"
                f"👁 {format_number(post_data.get('views', 0))} | ❤️ {format_number(post_data.get('reactions', 0))}"
            )
        else:
            await msg.edit_text(f"❌ Ошибка публикации: {queue_item.error_message or 'неизвестная ошибка'}")


# ============ ОЧИСТКА ОЧЕРЕДИ ============

async def clear_old_queue(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Удалить все pending посты из очереди (админ)."""
    if not await is_admin(update.effective_user.id):
        await update.message.reply_text("❌ Нет доступа")
        return
    
    msg = await update.message.reply_text("🧹 Очищаю очередь...")
    
    async with AsyncSessionLocal() as session:
        result = await session.execute(
            select(PostQueue).where(PostQueue.status == "pending")
        )
        items = result.scalars().all()
        
        deleted = len(items)
        for item in items:
            await session.delete(item)
        
        await session.commit()
    
    await msg.edit_text(f"✅ Удалено {deleted} постов из очереди")


async def clear_failed_queue(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Удалить все failed посты из очереди (админ)."""
    if not await is_admin(update.effective_user.id):
        await update.message.reply_text("❌ Нет доступа")
        return
    
    msg = await update.message.reply_text("🧹 Очищаю failed посты...")
    
    async with AsyncSessionLocal() as session:
        result = await session.execute(
            select(PostQueue).where(PostQueue.status == "failed")
        )
        items = result.scalars().all()
        
        deleted = len(items)
        for item in items:
            await session.delete(item)
        
        await session.commit()
    
    await msg.edit_text(f"✅ Удалено {deleted} failed постов")


async def clear_all_queue(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Удалить все посты из очереди (админ)."""
    if not await is_admin(update.effective_user.id):
        await update.message.reply_text("❌ Нет доступа")
        return
    
    msg = await update.message.reply_text("🧹 Очищаю всю очередь...")
    
    async with AsyncSessionLocal() as session:
        result = await session.execute(select(PostQueue))
        items = result.scalars().all()
        
        deleted = len(items)
        for item in items:
            await session.delete(item)
        
        await session.commit()
    
    await msg.edit_text(f"✅ Удалено {deleted} постов из очереди")


async def clear_project_queue(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Удалить все посты из очереди текущего проекта."""
    project = await require_project(update, context)
    
    if not project:
        return
    
    msg = await update.message.reply_text(f"🧹 Очищаю очередь проекта «{project.name}»...")
    
    async with AsyncSessionLocal() as session:
        result = await session.execute(
            select(PostQueue).where(PostQueue.project_id == project.id)
        )
        items = result.scalars().all()
        
        deleted = len(items)
        for item in items:
            await session.delete(item)
        
        await session.commit()
    
    await msg.edit_text(f"✅ Удалено {deleted} постов из очереди проекта «{project.name}»")


async def reset_history(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Сбрасывает историю parsed_posts для текущего проекта."""
    project = await require_project(update, context)
    
    if not project:
        return
    
    async with AsyncSessionLocal() as session:
        result = await session.execute(
            select(ParsedPost).where(ParsedPost.project_id == project.id)
        )
        items = result.scalars().all()
        
        deleted = len(items)
        for item in items:
            await session.delete(item)
        
        await session.commit()
    
    await update.message.reply_text(
        f"✅ История парсинга для проекта «{project.name}» сброшена.\n"
        f"Удалено {deleted} записей.\n"
        f"При следующем парсинге все посты будут считаться новыми."
    )


# ============ ТЕСТ ============

async def test_scraper(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Тестирует парсер на указанном канале."""
    if not context.args:
        await update.message.reply_text("ℹ️ /test [username]")
        return
    
    username = context.args[0].replace("@", "")
    msg = await update.message.reply_text(f"🔍 Тестирую @{username}...")
    
    async with TelegramScraper() as scraper:
        info = await scraper.get_channel_info(username)
        if not info:
            await msg.edit_text("❌ Канал не найден")
            return
        
        posts = await scraper.get_posts(username, limit=5)
        
        if posts:
            text = f"📨 @{username}\nНайдено: {len(posts)}\n\n"
            for p in posts[:5]:
                text += f"👁 {format_number(p['views'])} | ❤️ {format_number(p['reactions'])}\n"
                text += f"📎 {'📷' if p.get('media_type') == 'photo' else '🎬' if p.get('media_type') == 'video' else '📝'}\n\n"
        else:
            text = "❌ Посты не найдены"
    
    await msg.edit_text(text)


async def debug_reactions(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Отладка: показывает реакции на посты канала."""
    if not context.args:
        await update.message.reply_text("ℹ️ /debug_reactions [username]")
        return
    
    username = context.args[0].replace("@", "")
    msg = await update.message.reply_text(f"🔍 Анализирую реакции @{username}...")
    
    async with TelegramScraper() as scraper:
        posts = await scraper.get_posts(username, limit=10)
        
        if not posts:
            await msg.edit_text("❌ Посты не найдены")
            return
        
        text = f"📊 <b>Реакции @{username}</b>\n\n"
        for p in posts[:10]:
            text += f"👁 {format_number(p['views'])} | ❤️ {format_number(p.get('reactions', 0))}"
            if p.get('reactions_detail'):
                text += f" {p['reactions_detail']}"
            text += "\n"
    
    await msg.edit_text(text, parse_mode="HTML")


# ============ BROADCAST ============

async def broadcast_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Начинает рассылку сообщения всем пользователям (админ)."""
    if not await is_admin(update.effective_user.id):
        await update.message.reply_text("❌ Нет доступа")
        return ConversationHandler.END
    
    await update.message.reply_text(
        "📢 <b>Рассылка</b>\n\n"
        "Отправьте текст сообщения для всех пользователей:",
        parse_mode="HTML"
    )
    return AWAITING_BROADCAST_MESSAGE


async def broadcast_send(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Отправляет рассылку всем пользователям."""
    text = update.message.text
    
    async with AsyncSessionLocal() as session:
        result = await session.execute(select(User))
        users = result.scalars().all()
    
    sent = 0
    failed = 0
    
    for user in users:
        try:
            await context.bot.send_message(user.telegram_id, text)
            sent += 1
            await asyncio.sleep(0.05)
        except Exception as e:
            failed += 1
    
    await update.message.reply_text(
        f"✅ Рассылка завершена!\n\n"
        f"📤 Отправлено: {sent}\n"
        f"❌ Ошибок: {failed}"
    )
    
    return ConversationHandler.END


# ============ MEDIA FILTER ============

async def media_filter_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает выбор медиа-фильтра."""
    query = update.callback_query
    await query.answer()
    
    filter_type = query.data.replace("media_", "")
    
    if filter_type == "video":
        keyboard = [
            [InlineKeyboardButton("Без ограничений", callback_data="duration_none")],
            [InlineKeyboardButton("До 1 минуты", callback_data="duration_60")],
            [InlineKeyboardButton("До 3 минут", callback_data="duration_180")],
            [InlineKeyboardButton("До 5 минут", callback_data="duration_300")],
            [InlineKeyboardButton("До 10 минут", callback_data="duration_600")],
        ]
        await query.edit_message_text(
            "🎬 <b>Максимальная длительность видео:</b>",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode="HTML"
        )
        return AWAITING_MEDIA_FILTER
    else:
        context.user_data['temp_source_media'] = filter_type
        context.user_data['temp_source_duration'] = None
        await query.edit_message_text(f"✅ Медиа-фильтр: {filter_type}")
        return await ask_remove_text(query)


async def duration_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает выбор длительности видео."""
    query = update.callback_query
    await query.answer()
    
    duration = query.data.replace("duration_", "")
    context.user_data['temp_source_media'] = "video"
    context.user_data['temp_source_duration'] = None if duration == "none" else int(duration)
    
    duration_text = "без ограничений" if duration == "none" else f"до {int(duration)//60} мин"
    await query.edit_message_text(f"✅ Видео: {duration_text}")
    return await ask_remove_text(query)


async def ask_remove_text(query):
    """Спрашивает, удалять ли текст оригинала."""
    keyboard = [
        [InlineKeyboardButton("✅ Да, удалять текст", callback_data="text_yes")],
        [InlineKeyboardButton("❌ Нет, оставлять", callback_data="text_no")],
    ]
    await query.message.reply_text(
        "📝 <b>Удалять текст оригинального поста?</b>",
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode="HTML"
    )
    return AWAITING_REMOVE_TEXT


async def remove_text_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает выбор удаления текста."""
    query = update.callback_query
    await query.answer()
    
    remove = query.data == "text_yes"
    context.user_data['temp_remove_text'] = remove
    
    temp = context.user_data.get('temp_source')
    if temp:
        temp['media_filter'] = context.user_data.get('temp_source_media', 'all')
        temp['max_video_duration'] = context.user_data.get('temp_source_duration')
        temp['remove_original_text'] = remove
        criteria = context.user_data.get('temp_criteria', {})
        await save_source_with_criteria(query, context, temp, criteria)
    
    return ConversationHandler.END


# ============ АДМИН ТАРИФЫ ============

async def admin_set_tariff_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Установка тарифа пользователю (админ)."""
    if not await is_admin(update.effective_user.id):
        await update.message.reply_text("❌ Нет доступа")
        return
    
    if not context.args:
        await update.message.reply_text("ℹ️ /admin_set_tariff [user_id] [trial|basic|pro|unlimited]")
        return
    
    try:
        user_id = int(context.args[0])
        tariff = context.args[1] if len(context.args) > 1 else "basic"
    except:
        await update.message.reply_text("❌ Неверный формат. Пример: /admin_set_tariff 123456 pro")
        return
    
    async with AsyncSessionLocal() as session:
        result = await session.execute(select(User).where(User.telegram_id == user_id))
        user = result.scalar_one_or_none()
        
        if not user:
            await update.message.reply_text("❌ Пользователь не найден")
            return
        
        if tariff == "trial":
            user.tariff = "trial"
            user.trial_ends_at = datetime.utcnow() + timedelta(days=5)
            user.subscription_active = False
        elif tariff == "basic":
            user.tariff = "basic"
            user.subscription_active = True
            user.subscription_ends_at = datetime.utcnow() + timedelta(days=30)
        elif tariff == "pro":
            user.tariff = "pro"
            user.subscription_active = True
            user.subscription_ends_at = datetime.utcnow() + timedelta(days=30)
        elif tariff == "unlimited":
            user.tariff = "unlimited"
            user.subscription_active = True
            user.subscription_ends_at = datetime.utcnow() + timedelta(days=365)
        
        await session.commit()
    
    await update.message.reply_text(
        f"✅ Тариф пользователя {user_id} установлен: {tariff}"
    )


async def admin_extend_trial_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Продлевает триал пользователю (админ)."""
    if not await is_admin(update.effective_user.id):
        await update.message.reply_text("❌ Нет доступа")
        return
    
    if not context.args:
        await update.message.reply_text("ℹ️ /admin_extend_trial [user_id] [days]")
        return
    
    try:
        user_id = int(context.args[0])
        days = int(context.args[1]) if len(context.args) > 1 else 5
    except:
        await update.message.reply_text("❌ Неверный формат. Пример: /admin_extend_trial 123456 7")
        return
    
    async with AsyncSessionLocal() as session:
        result = await session.execute(select(User).where(User.telegram_id == user_id))
        user = result.scalar_one_or_none()
        
        if not user:
            await update.message.reply_text("❌ Пользователь не найден")
            return
        
        if user.trial_ends_at and user.trial_ends_at > datetime.utcnow():
            user.trial_ends_at = user.trial_ends_at + timedelta(days=days)
        else:
            user.trial_ends_at = datetime.utcnow() + timedelta(days=days)
        
        user.tariff = "trial"
        await session.commit()
    
    await update.message.reply_text(
        f"✅ Триал пользователя {user_id} продлён на {days} дней до {user.trial_ends_at.strftime('%d.%m.%Y')}"
    )


# ============ ОТМЕНА ============

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Отменяет текущее действие."""
    context.user_data.clear()
    await update.message.reply_text("❌ Действие отменено")
    return ConversationHandler.END