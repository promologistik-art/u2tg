import os
import logging
import asyncio
from datetime import datetime, timedelta
from io import BytesIO
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ContextTypes, ConversationHandler
from sqlalchemy import select, func, delete, update as sql_update
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from config import Config
from database import AsyncSessionLocal
from models import User, Project, PostQueue
from backup import BackupService
from .utils import is_admin, get_user_projects_count, TARIFF_LIMITS
from .constants import AWAITING_TARIFF_SELECT, AWAITING_BROADCAST_MESSAGE

logger = logging.getLogger(__name__)


# ============ ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ============

async def update_user_limits_direct(user: User, tariff: str):
    """Обновляет лимиты пользователя в соответствии с тарифом."""
    limits = TARIFF_LIMITS.get(tariff, TARIFF_LIMITS["trial"])
    user.max_projects = limits["max_projects"]
    user.max_sources_per_project = limits["max_sources_per_project"]
    user.min_post_interval_minutes = limits["min_post_interval"]
    user.min_check_interval_minutes = limits["min_check_interval"]


# ============ АДМИН-ПАНЕЛЬ ============

async def admin_panel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await is_admin(update.effective_user.id):
        await update.message.reply_text("❌ Нет доступа")
        return
    
    source_state = "ВКЛ ✅" if Config.SHOW_SOURCE_SIGNATURE else "ВЫКЛ ❌"
    
    keyboard = [
        [InlineKeyboardButton("👥 Пользователи", callback_data="admin_users_list")],
        [InlineKeyboardButton("💳 Управление тарифами", callback_data="admin_tariff_menu")],
        [InlineKeyboardButton(f"🔧 Источник: {source_state}", callback_data="admin_toggle_source")],
        [InlineKeyboardButton("💾 Создать бэкап", callback_data="admin_backup_create")],
        [InlineKeyboardButton("📦 Список бэкапов", callback_data="admin_backup_list")],
        [InlineKeyboardButton("📊 Экспорт в Excel", callback_data="admin_export")],
        [InlineKeyboardButton("📈 Отчёт по пользователям", callback_data="admin_report")],
        [InlineKeyboardButton("🔍 Диагностика", callback_data="admin_diagnose")],
        [InlineKeyboardButton("🧹 Очистить очередь", callback_data="admin_clear_queue")],
        [InlineKeyboardButton("🗑️ Очистить failed", callback_data="admin_clear_failed")],
        [InlineKeyboardButton("📢 Рассылка", callback_data="admin_broadcast")],
    ]
    
    if update.callback_query:
        await update.callback_query.edit_message_text(
            "👑 <b>Админ-панель</b>\n\nВыберите действие:",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode="HTML"
        )
    else:
        await update.message.reply_text(
            "👑 <b>Админ-панель</b>\n\nВыберите действие:",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode="HTML"
        )


async def admin_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    if not await is_admin(update.effective_user.id):
        await query.edit_message_text("❌ Нет доступа")
        return
    
    action = query.data
    
    if action == "admin_back":
        return
    
    if action == "admin_users_list":
        await show_admin_users(query)
    elif action == "admin_tariff_menu":
        await show_tariff_menu(query)
    elif action == "admin_toggle_source":
        await toggle_source_signature(query)
    elif action == "admin_backup_create":
        await create_backup_admin(query)
    elif action == "admin_backup_list":
        await list_backups_admin(query)
    elif action == "admin_export":
        await export_users_excel(query, context)
    elif action == "admin_report":
        await send_daily_report(query, context)
    elif action == "admin_diagnose":
        await show_diagnose_admin(query)
    elif action == "admin_clear_queue":
        await clear_queue_admin(query)
    elif action == "admin_clear_failed":
        await clear_failed_admin(query)
    elif action == "admin_broadcast":
        await broadcast_start(update, context)
    elif action.startswith("tariff_set_"):
        tariff = action.replace("tariff_set_", "")
        await tariff_select_user(query, tariff, context)
    elif action.startswith("user_tariff_"):
        user_id = int(action.replace("user_tariff_", ""))
        tariff = context.user_data.get('selected_tariff')
        await confirm_set_tariff(query, user_id, tariff)
    elif action.startswith("extend_user_"):
        user_id = int(action.replace("extend_user_", ""))
        await extend_trial_days(query, user_id)
    elif action.startswith("deactivate_user_"):
        user_id = int(action.replace("deactivate_user_", ""))
        await deactivate_user(query, user_id)
    elif action.startswith("activate_user_"):
        user_id = int(action.replace("activate_user_", ""))
        await activate_user(query, user_id)
    elif action.startswith("user_manage_"):
        user_id = int(action.replace("user_manage_", ""))
        await show_user_manage_menu(query, user_id)
    elif action == "admin_set_tariff":
        await tariff_select_menu(query)
    elif action == "admin_extend_trial":
        await extend_trial_start(query)
    elif action == "admin_deactivate":
        await deactivate_menu(query)
    elif action == "admin_activate":
        await activate_menu(query)
    elif action.startswith("tariff_for_"):
        user_id = int(action.replace("tariff_for_", ""))
        await tariff_select_menu_for_user(query, user_id)
    elif action.startswith("set_tariff_"):
        parts = action.split("_")
        user_id = int(parts[2])
        tariff = parts[3]
        await confirm_set_tariff(query, user_id, tariff)


async def toggle_source_signature(query):
    new_state = Config.toggle_source_signature()
    state_text = "ВКЛ ✅" if new_state else "ВЫКЛ ❌"
    keyboard = [[InlineKeyboardButton("◀️ Назад", callback_data="admin_back")]]
    await query.edit_message_text(
        f"🔧 <b>Отображение источника</b>\n\nТекущее состояние: {state_text}",
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode="HTML"
    )


async def show_admin_users(query):
    async with AsyncSessionLocal() as session:
        result = await session.execute(select(User).order_by(User.created_at.desc()).limit(20))
        users = result.scalars().all()
    
    text = f"👥 <b>Пользователи ({len(users)}):</b>\n\n"
    keyboard = []
    
    for u in users:
        projects_count = await get_user_projects_count(u.telegram_id)
        
        if u.is_admin:
            status_icon = "👑"
            tariff_display = "Безлимит"
        else:
            if u.subscription_active:
                status_icon = "💎"
            elif u.trial_ends_at and u.trial_ends_at > datetime.utcnow():
                status_icon = "🎁"
            else:
                status_icon = "🔴"
            tariff_display = TARIFF_LIMITS.get(u.tariff, {}).get('name', '—')
        
        display_name = u.full_name or u.username or "Пользователь"
        
        text += f"{status_icon} {display_name}"
        if u.username:
            text += f" (@{u.username})\n"
        else:
            text += "\n"
        text += f"  🆔 {u.telegram_id} | 📁 {projects_count} проектов\n"
        text += f"  💳 {tariff_display}\n"
        
        if u.subscription_active and u.subscription_ends_at:
            text += f"  📅 Подписка до: {u.subscription_ends_at.strftime('%d.%m.%Y')}\n"
        elif u.trial_ends_at and u.trial_ends_at > datetime.utcnow():
            text += f"  📅 Триал до: {u.trial_ends_at.strftime('%d.%m.%Y')}\n"
        text += "\n"
        
        keyboard.append([InlineKeyboardButton(f"⚙️ Управлять {display_name[:15]}", callback_data=f"user_manage_{u.telegram_id}")])
    
    keyboard.append([InlineKeyboardButton("◀️ Назад", callback_data="admin_back")])
    await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode="HTML")


# ============ УПРАВЛЕНИЕ ТАРИФАМИ ============

async def show_tariff_menu(query):
    text = "💳 <b>Управление тарифами</b>\n\nВыберите действие:"
    keyboard = [
        [InlineKeyboardButton("💎 Установить тариф", callback_data="admin_set_tariff")],
        [InlineKeyboardButton("🎁 Продлить триал", callback_data="admin_extend_trial")],
        [InlineKeyboardButton("❌ Деактивировать", callback_data="admin_deactivate")],
        [InlineKeyboardButton("✅ Активировать", callback_data="admin_activate")],
        [InlineKeyboardButton("◀️ Назад", callback_data="admin_back")],
    ]
    await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode="HTML")


async def admin_set_tariff_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await is_admin(update.effective_user.id):
        await update.message.reply_text("❌ Нет доступа")
        return ConversationHandler.END
    
    text = (
        "💎 <b>Выберите тариф:</b>\n\n"
        "🟡 <b>Базовый</b> — 290 ₽/мес\n   • 1 проект, 3 источника, постинг от 2ч\n\n"
        "🟠 <b>Стандарт</b> — 590 ₽/мес\n   • 3 проекта, 5 источников, постинг от 1ч\n\n"
        "🔴 <b>PRO</b> — 990 ₽/мес\n   • 10 проектов, 10 источников, постинг от 30мин\n\n"
        "👑 <b>Безлимит</b> — 1990 ₽/мес\n   • Без ограничений"
    )
    keyboard = [
        [InlineKeyboardButton("🟡 Базовый", callback_data="tariff_set_basic")],
        [InlineKeyboardButton("🟠 Стандарт", callback_data="tariff_set_standard")],
        [InlineKeyboardButton("🔴 PRO", callback_data="tariff_set_pro")],
        [InlineKeyboardButton("👑 Безлимит", callback_data="tariff_set_unlimited")],
    ]
    await update.message.reply_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode="HTML")
    return AWAITING_TARIFF_SELECT


async def tariff_select_user(query, tariff: str, context):
    context.user_data['selected_tariff'] = tariff
    async with AsyncSessionLocal() as session:
        result = await session.execute(select(User).order_by(User.created_at.desc()).limit(30))
        users = result.scalars().all()
    
    text = f"💎 <b>Выбран тариф: {TARIFF_LIMITS.get(tariff, {}).get('name', tariff)}</b>\n\nВыберите пользователя:"
    keyboard = []
    for u in users:
        display_name = u.full_name or u.username or f"ID:{u.telegram_id}"
        keyboard.append([InlineKeyboardButton(display_name[:30], callback_data=f"user_tariff_{u.telegram_id}")])
    keyboard.append([InlineKeyboardButton("◀️ Назад", callback_data="admin_tariff_menu")])
    await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode="HTML")


async def confirm_set_tariff(query, user_id: int, tariff: str):
    async with AsyncSessionLocal() as session:
        result = await session.execute(select(User).where(User.telegram_id == user_id))
        user = result.scalar_one_or_none()
        if not user:
            await query.edit_message_text("❌ Пользователь не найден")
            return
        
        user.subscription_active = True
        user.subscription_ends_at = datetime.utcnow() + timedelta(days=30)
        user.tariff = tariff
        user.trial_ends_at = None
        
        await update_user_limits_direct(user, tariff)
        
        await session.commit()
        
        tariff_name = TARIFF_LIMITS.get(tariff, {}).get('name', tariff)
        try:
            await query.message.bot.send_message(
                chat_id=user_id,
                text=f"✅ <b>Тариф активирован!</b>\n💎 {tariff_name}\n📅 Подписка до: {user.subscription_ends_at.strftime('%d.%m.%Y')}\n\n"
                     f"📊 Ваши лимиты:\n"
                     f"• Проектов: {user.max_projects}\n"
                     f"• Источников на проект: {user.max_sources_per_project}\n"
                     f"• Мин. интервал постинга: {user.min_post_interval_minutes} мин\n"
                     f"• Мин. интервал парсинга: {user.min_check_interval_minutes} мин",
                parse_mode="HTML"
            )
        except Exception as e:
            logger.error(f"Failed to notify user {user_id}: {e}")
    
    keyboard = [[InlineKeyboardButton("◀️ В админ-панель", callback_data="admin_back")]]
    await query.edit_message_text(
        f"✅ Тариф <b>{tariff_name}</b> подключен!\n👤 @{user.username or user.telegram_id}\n📅 Подписка до: {user.subscription_ends_at.strftime('%d.%m.%Y')}",
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode="HTML"
    )


async def admin_extend_trial_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await is_admin(update.effective_user.id):
        await update.message.reply_text("❌ Нет доступа")
        return ConversationHandler.END
    
    async with AsyncSessionLocal() as session:
        result = await session.execute(
            select(User).where(
                User.subscription_active == False,
                User.trial_ends_at > datetime.utcnow()
            ).order_by(User.trial_ends_at.asc()).limit(20)
        )
        users = result.scalars().all()
    
    if not users:
        await update.message.reply_text("📭 Нет пользователей на активном триале")
        return ConversationHandler.END
    
    text = "🎁 <b>Продлить триал</b>\n\nВыберите пользователя:"
    keyboard = []
    for u in users:
        days_left = (u.trial_ends_at - datetime.utcnow()).days if u.trial_ends_at else 0
        display_name = u.full_name or u.username or f"ID:{u.telegram_id}"
        keyboard.append([InlineKeyboardButton(f"➕ {display_name[:20]} ({days_left} дн)", callback_data=f"extend_user_{u.telegram_id}")])
    keyboard.append([InlineKeyboardButton("◀️ Назад", callback_data="admin_back")])
    await update.message.reply_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode="HTML")
    return ConversationHandler.END


async def extend_trial_days(query, user_id: int):
    async with AsyncSessionLocal() as session:
        result = await session.execute(select(User).where(User.telegram_id == user_id))
        user = result.scalar_one_or_none()
        if not user:
            await query.edit_message_text("❌ Пользователь не найден")
            return
        
        if user.trial_ends_at and user.trial_ends_at > datetime.utcnow():
            user.trial_ends_at = user.trial_ends_at + timedelta(days=7)
        else:
            user.trial_ends_at = datetime.utcnow() + timedelta(days=7)
        
        user.tariff = "trial"
        await update_user_limits_direct(user, "trial")
        user.subscription_active = False
        user.subscription_ends_at = None
        
        await session.commit()
        
        try:
            await query.message.bot.send_message(
                chat_id=user_id,
                text=f"🎁 <b>Триал продлён!</b>\n📅 Триал до: {user.trial_ends_at.strftime('%d.%m.%Y')}\n\n"
                     f"📊 Ваши лимиты:\n"
                     f"• Проектов: {user.max_projects}\n"
                     f"• Источников на проект: {user.max_sources_per_project}\n"
                     f"• Мин. интервал постинга: {user.min_post_interval_minutes} мин",
                parse_mode="HTML"
            )
        except:
            pass
    
    keyboard = [[InlineKeyboardButton("◀️ В админ-панель", callback_data="admin_back")]]
    await query.edit_message_text(
        f"✅ Триал продлён!\n👤 @{user.username or user.telegram_id}\n📅 Триал до: {user.trial_ends_at.strftime('%d.%m.%Y')}",
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode="HTML"
    )


async def extend_trial_start(query):
    await admin_extend_trial_start(query, None)


async def deactivate_user(query, user_id: int):
    async with AsyncSessionLocal() as session:
        result = await session.execute(select(User).where(User.telegram_id == user_id))
        user = result.scalar_one_or_none()
        if not user:
            await query.edit_message_text("❌ Пользователь не найден")
            return
        user.subscription_active = False
        user.trial_ends_at = datetime.utcnow() - timedelta(days=1)
        user.subscription_ends_at = None
        await session.commit()
    
    keyboard = [[InlineKeyboardButton("◀️ В админ-панель", callback_data="admin_back")]]
    await query.edit_message_text(f"✅ Пользователь деактивирован", reply_markup=InlineKeyboardMarkup(keyboard))


async def activate_user(query, user_id: int):
    async with AsyncSessionLocal() as session:
        result = await session.execute(select(User).where(User.telegram_id == user_id))
        user = result.scalar_one_or_none()
        if not user:
            await query.edit_message_text("❌ Пользователь не найден")
            return
        user.trial_ends_at = datetime.utcnow() + timedelta(days=5)
        user.tariff = "trial"
        await update_user_limits_direct(user, "trial")
        user.subscription_active = False
        user.subscription_ends_at = None
        await session.commit()
    
    keyboard = [[InlineKeyboardButton("◀️ В админ-панель", callback_data="admin_back")]]
    await query.edit_message_text(f"✅ Пользователь активирован (триал 5 дней)", reply_markup=InlineKeyboardMarkup(keyboard))


# ============ МЕНЮ УПРАВЛЕНИЯ ПОЛЬЗОВАТЕЛЕМ ============

async def show_user_manage_menu(query, user_id: int):
    async with AsyncSessionLocal() as session:
        result = await session.execute(select(User).where(User.telegram_id == user_id))
        user = result.scalar_one_or_none()
    if not user:
        await query.edit_message_text("❌ Пользователь не найден")
        return
    
    tariff_name = TARIFF_LIMITS.get(user.tariff, {}).get('name', user.tariff)
    display_name = user.full_name or user.username or f"ID:{user.telegram_id}"
    
    text = f"⚙️ <b>{display_name}</b>\n💳 {tariff_name}\n"
    
    if user.subscription_active and user.subscription_ends_at:
        text += f"📅 Подписка до: {user.subscription_ends_at.strftime('%d.%m.%Y %H:%M')}\n"
    elif user.trial_ends_at and user.trial_ends_at > datetime.utcnow():
        text += f"📅 Триал до: {user.trial_ends_at.strftime('%d.%m.%Y %H:%M')}\n"
    
    text += f"\n📊 Лимиты:\n"
    text += f"• Проектов: {user.max_projects}\n"
    text += f"• Источников на проект: {user.max_sources_per_project}\n"
    text += f"• Мин. интервал постинга: {user.min_post_interval_minutes} мин\n"
    text += f"• Мин. интервал парсинга: {user.min_check_interval_minutes} мин\n\n"
    text += f"Выберите действие:"
    
    keyboard = [
        [InlineKeyboardButton("💎 Установить тариф", callback_data=f"tariff_for_{user_id}")],
    ]
    
    if user.subscription_active:
        keyboard.append([InlineKeyboardButton("🎁 Продлить триал", callback_data=f"extend_user_{user_id}")])
        keyboard.append([InlineKeyboardButton("❌ Деактивировать", callback_data=f"deactivate_user_{user_id}")])
    else:
        if user.trial_ends_at and user.trial_ends_at > datetime.utcnow():
            keyboard.append([InlineKeyboardButton("🎁 Продлить триал", callback_data=f"extend_user_{user_id}")])
        keyboard.append([InlineKeyboardButton("✅ Активировать (триал 5дн)", callback_data=f"activate_user_{user_id}")])
    
    keyboard.append([InlineKeyboardButton("◀️ Назад к списку", callback_data="admin_users_list")])
    
    await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode="HTML")


async def tariff_select_menu(query):
    text = "💎 <b>Выберите тариф:</b>\n\n🟡 Базовый 290₽ | 🟠 Стандарт 590₽ | 🔴 PRO 990₽ | 👑 Безлимит 1990₽"
    keyboard = [
        [InlineKeyboardButton("🟡 Базовый", callback_data="tariff_set_basic")],
        [InlineKeyboardButton("🟠 Стандарт", callback_data="tariff_set_standard")],
        [InlineKeyboardButton("🔴 PRO", callback_data="tariff_set_pro")],
        [InlineKeyboardButton("👑 Безлимит", callback_data="tariff_set_unlimited")],
        [InlineKeyboardButton("◀️ Назад", callback_data="admin_tariff_menu")],
    ]
    await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode="HTML")


async def tariff_select_menu_for_user(query, user_id: int):
    keyboard = [
        [InlineKeyboardButton("🟡 Базовый", callback_data=f"set_tariff_{user_id}_basic")],
        [InlineKeyboardButton("🟠 Стандарт", callback_data=f"set_tariff_{user_id}_standard")],
        [InlineKeyboardButton("🔴 PRO", callback_data=f"set_tariff_{user_id}_pro")],
        [InlineKeyboardButton("👑 Безлимит", callback_data=f"set_tariff_{user_id}_unlimited")],
        [InlineKeyboardButton("◀️ Назад", callback_data=f"user_manage_{user_id}")],
    ]
    await query.edit_message_text(f"💎 <b>Выберите тариф для пользователя:</b>", reply_markup=InlineKeyboardMarkup(keyboard), parse_mode="HTML")


async def deactivate_menu(query):
    async with AsyncSessionLocal() as session:
        result = await session.execute(select(User).where(User.subscription_active == True).order_by(User.created_at.desc()).limit(20))
        users = result.scalars().all()
    if not users:
        await query.edit_message_text("📭 Нет активных пользователей", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("◀️ Назад", callback_data="admin_tariff_menu")]]))
        return
    text = "❌ <b>Выберите пользователя для деактивации:</b>"
    keyboard = []
    for u in users:
        display_name = u.full_name or u.username or f"ID:{u.telegram_id}"
        keyboard.append([InlineKeyboardButton(f"❌ {display_name[:20]}", callback_data=f"deactivate_user_{u.telegram_id}")])
    keyboard.append([InlineKeyboardButton("◀️ Назад", callback_data="admin_tariff_menu")])
    await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode="HTML")


async def activate_menu(query):
    async with AsyncSessionLocal() as session:
        result = await session.execute(
            select(User).where(
                User.subscription_active == False,
                User.trial_ends_at < datetime.utcnow()
            ).order_by(User.created_at.desc()).limit(20)
        )
        users = result.scalars().all()
    if not users:
        await query.edit_message_text("📭 Нет неактивных пользователей (с истекшим триалом)", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("◀️ Назад", callback_data="admin_tariff_menu")]]))
        return
    text = "✅ <b>Выберите пользователя для активации (триал 5 дней):</b>"
    keyboard = []
    for u in users:
        display_name = u.full_name or u.username or f"ID:{u.telegram_id}"
        keyboard.append([InlineKeyboardButton(f"✅ {display_name[:20]}", callback_data=f"activate_user_{u.telegram_id}")])
    keyboard.append([InlineKeyboardButton("◀️ Назад", callback_data="admin_tariff_menu")])
    await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode="HTML")


# ============ БЭКАПЫ ============

async def create_backup_admin(query):
    backup_service = BackupService()
    backup_path = backup_service.create_backup()
    if backup_path:
        try:
            with open(backup_path, 'rb') as f:
                await query.message.reply_document(document=f, filename=os.path.basename(backup_path), caption="✅ Бэкап создан")
        except:
            pass
        await query.edit_message_text("✅ Бэкап создан и отправлен!", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("◀️ Назад", callback_data="admin_back")]]))
    else:
        await query.edit_message_text("❌ Ошибка", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("◀️ Назад", callback_data="admin_back")]]))


async def list_backups_admin(query):
    backup_service = BackupService()
    backups = backup_service.list_backups()
    if not backups:
        text = "📭 Бэкапов нет"
    else:
        text = "📦 <b>Бэкапы:</b>\n\n"
        for b in backups[:10]:
            text += f"• {b['name']}\n  📅 {b['created']} | 📦 {b['size_mb']} MB\n\n"
    await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("◀️ Назад", callback_data="admin_back")]]), parse_mode="HTML")


# ============ ЭКСПОРТ В EXCEL ============

async def export_users_excel(query, context):
    async with AsyncSessionLocal() as session:
        result = await session.execute(select(User).order_by(User.created_at.desc()))
        users = result.scalars().all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Пользователи"
    
    headers = ["Telegram ID", "Username", "Full Name", "Admin", "Tariff", "Status", "Subscription End", "Trial End", "Projects", "Sources Limit", "Post Interval", "Parse Interval", "Parsed Today", "Posted Today", "Created At"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    for row, u in enumerate(users, 2):
        projects_count = await get_user_projects_count(u.telegram_id)
        
        if u.is_admin:
            tariff_name = "Безлимит"
            status = "Админ"
        else:
            tariff_name = TARIFF_LIMITS.get(u.tariff, {}).get('name', u.tariff)
            if u.subscription_active:
                status = "Подписка"
            elif u.trial_ends_at and u.trial_ends_at > datetime.utcnow():
                status = "Триал"
            else:
                status = "Нет доступа"
        
        ws.cell(row=row, column=1, value=u.telegram_id)
        ws.cell(row=row, column=2, value=u.username or "")
        ws.cell(row=row, column=3, value=u.full_name or "")
        ws.cell(row=row, column=4, value="Да" if u.is_admin else "Нет")
        ws.cell(row=row, column=5, value=tariff_name)
        ws.cell(row=row, column=6, value=status)
        ws.cell(row=row, column=7, value=u.subscription_ends_at.strftime("%d.%m.%Y") if u.subscription_ends_at else "")
        ws.cell(row=row, column=8, value=u.trial_ends_at.strftime("%d.%m.%Y") if u.trial_ends_at else "")
        ws.cell(row=row, column=9, value=projects_count)
        ws.cell(row=row, column=10, value=u.max_sources_per_project)
        ws.cell(row=row, column=11, value=u.min_post_interval_minutes)
        ws.cell(row=row, column=12, value=u.min_check_interval_minutes)
        ws.cell(row=row, column=13, value=u.posts_parsed_today)
        ws.cell(row=row, column=14, value=u.posts_posted_today)
        ws.cell(row=row, column=15, value=u.created_at.strftime("%d.%m.%Y %H:%M") if u.created_at else "")
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    await context.bot.send_document(chat_id=query.message.chat_id, document=output, filename=f"users_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx", caption="📊 Экспорт пользователей")
    await query.edit_message_text("✅ Отчёт отправлен!", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("◀️ Назад", callback_data="admin_back")]]))


async def send_daily_report(query, context):
    async with AsyncSessionLocal() as session:
        result = await session.execute(select(User).order_by(User.created_at.desc()))
        users = result.scalars().all()
    
    now = datetime.utcnow()
    total_users = len(users)
    new_today = sum(1 for u in users if u.created_at and (now - u.created_at).days < 1)
    on_trial = sum(1 for u in users if not u.subscription_active and u.trial_ends_at and u.trial_ends_at > now)
    paid = sum(1 for u in users if u.subscription_active)
    
    text = f"📊 <b>Ежедневный отчёт</b>\n📅 {now.strftime('%d.%m.%Y')}\n\n👥 Всего: {total_users}\n🆕 Новых: {new_today}\n🎁 На триале: {on_trial}\n💎 Платных: {paid}"
    await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("◀️ Назад", callback_data="admin_back")]]), parse_mode="HTML")


# ============ ДИАГНОСТИКА И ОЧИСТКА ============

async def show_diagnose_admin(query):
    async with AsyncSessionLocal() as session:
        result = await session.execute(select(func.count()).select_from(User))
        users_count = result.scalar()
        result = await session.execute(select(func.count()).select_from(Project))
        projects_count = result.scalar()
        result = await session.execute(select(PostQueue).where(PostQueue.status == "pending"))
        pending = len(result.scalars().all())
    
    text = f"🔍 <b>Диагностика</b>\n\n👥 Пользователей: {users_count}\n📁 Проектов: {projects_count}\n📬 В очереди: {pending}"
    await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("◀️ Назад", callback_data="admin_back")]]), parse_mode="HTML")


async def clear_queue_admin(query):
    async with AsyncSessionLocal() as session:
        result = await session.execute(select(PostQueue).where(PostQueue.status == "pending"))
        items = result.scalars().all()
        deleted = len(items)
        for item in items:
            await session.delete(item)
        await session.commit()
    await query.edit_message_text(f"✅ Удалено {deleted} постов", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("◀️ Назад", callback_data="admin_back")]]))


async def clear_failed_admin(query):
    async with AsyncSessionLocal() as session:
        await session.execute(delete(PostQueue).where(PostQueue.status == "failed"))
        await session.commit()
    await query.edit_message_text("✅ Failed посты удалены", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("◀️ Назад", callback_data="admin_back")]]))


# ============ РАССЫЛКА ============

async def broadcast_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if hasattr(update, 'callback_query') and update.callback_query:
        query = update.callback_query
        await query.answer()
        if not await is_admin(update.effective_user.id):
            await query.edit_message_text("❌ Нет доступа")
            return ConversationHandler.END
        await query.edit_message_text("📢 <b>Рассылка</b>\n\nОтправьте текст сообщения.\n/cancel — отмена", parse_mode="HTML")
    else:
        if not await is_admin(update.effective_user.id):
            await update.message.reply_text("❌ Нет доступа")
            return ConversationHandler.END
        await update.message.reply_text("📢 <b>Рассылка</b>\n\nОтправьте текст сообщения.\n/cancel — отмена", parse_mode="HTML")
    context.user_data['awaiting_broadcast'] = True
    return AWAITING_BROADCAST_MESSAGE


async def broadcast_send(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.user_data.get('awaiting_broadcast'):
        return
    if not await is_admin(update.effective_user.id):
        await update.message.reply_text("❌ Нет доступа")
        return ConversationHandler.END
    
    text = update.message.text
    async with AsyncSessionLocal() as session:
        result = await session.execute(select(User))
        users = result.scalars().all()
    
    sent = 0
    msg = await update.message.reply_text(f"📢 Отправляю ({len(users)} пользователей)...")
    for user in users:
        try:
            await context.bot.send_message(chat_id=user.telegram_id, text=f"📢 <b>Сообщение от администратора:</b>\n\n{text}", parse_mode="HTML")
            sent += 1
        except:
            pass
        await asyncio.sleep(0.5)
    
    context.user_data['awaiting_broadcast'] = False
    await msg.edit_text(f"✅ Рассылка завершена!\n📨 Отправлено: {sent}")
    return ConversationHandler.END


async def admin_back_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    source_state = "ВКЛ ✅" if Config.SHOW_SOURCE_SIGNATURE else "ВЫКЛ ❌"
    
    keyboard = [
        [InlineKeyboardButton("👥 Пользователи", callback_data="admin_users_list")],
        [InlineKeyboardButton("💳 Управление тарифами", callback_data="admin_tariff_menu")],
        [InlineKeyboardButton(f"🔧 Источник: {source_state}", callback_data="admin_toggle_source")],
        [InlineKeyboardButton("💾 Создать бэкап", callback_data="admin_backup_create")],
        [InlineKeyboardButton("📦 Список бэкапов", callback_data="admin_backup_list")],
        [InlineKeyboardButton("📊 Экспорт в Excel", callback_data="admin_export")],
        [InlineKeyboardButton("📈 Отчёт по пользователям", callback_data="admin_report")],
        [InlineKeyboardButton("🔍 Диагностика", callback_data="admin_diagnose")],
        [InlineKeyboardButton("🧹 Очистить очередь", callback_data="admin_clear_queue")],
        [InlineKeyboardButton("🗑️ Очистить failed", callback_data="admin_clear_failed")],
        [InlineKeyboardButton("📢 Рассылка", callback_data="admin_broadcast")],
    ]
    
    await query.edit_message_text(
        "👑 <b>Админ-панель</b>\n\nВыберите действие:",
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode="HTML"
    )